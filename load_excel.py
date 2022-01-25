import openpyxl as xl
wb = xl.load_workbook(r"/Users/ankanasaumya/Documents/GitHub/transactions.xlsx") #Selecting Excel file
sheet = wb['Sheet1'] #Case Sensitive
cell = sheet['a1']
print(sheet.max_row)

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.90
    # print(corrected_price)
    corrected_price_cell = sheet.cell(row, 4)
    # print(row)
    corrected_price_cell = corrected_price
    # print(corrected_price)
    wb.save('trans.xlsx')